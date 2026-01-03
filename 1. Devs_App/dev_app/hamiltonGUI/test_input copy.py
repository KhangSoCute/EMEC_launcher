import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import os

# Read data from an Excel (.xlsx) file
filename = 'input_template.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb.active

# Export data to another Excel (.xlsx) file
new_filename = 'example2.xlsx'
new_wb = Workbook()
new_ws = new_wb.active

for row in ws.iter_rows():
    row_data = [cell.value for cell in row]
    new_ws.append(row_data)

new_wb.save(new_filename)

# Convert the new Excel file to a different format (.xls)
import win32com.client as win32

excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(os.path.abspath(new_filename))
new_filename_xls = 'input_test.xls'
wb.SaveAs(os.path.abspath(new_filename_xls), FileFormat=56)
wb.Close(True)
excel.Quit()