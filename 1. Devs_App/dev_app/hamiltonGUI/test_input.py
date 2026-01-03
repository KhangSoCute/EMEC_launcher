import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import os

# Export data to another Excel (.xlsx) file
new_filename = 'example2.xlsx'
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "Master"

new_ws.cell(row=2,column=2,value="No.Sample")

new_wb.save(new_filename)

# Convert the new Excel file to a different format (.xls)
import win32com.client as win32

excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(os.path.abspath(new_filename))
new_filename_xls = 'input_test.xls'
os.remove(new_filename_xls)
wb.SaveAs(os.path.abspath(new_filename_xls), FileFormat=56)
wb.Close(True)
excel.Quit()